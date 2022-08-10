# Heatmap Generation from MicroOrganism Data
# By William Pearson and Hannah Eccleston

#Import Packages
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import hashlib
import graphviz


###Define Class###
class SampleInformation:
    def __init__(self, workBookName, workSheetName, sampleName):
        self.workBookName = workBookName
        self.workSheetName = workSheetName
        self.sampleName = sampleName
        self.nameOveride = None
        self.iDColumnStart = None
        self.iDColumnEnd = None
        self.relativeAbundanceColumn = None
        self.dataRowStart = None
        self.dataRowEnd = None
        self.MicroOrganisms = None
        self.Parent = None
        self.microOrganismTree = []


    def SetUpMicroOrganismDatSet(self, iDColumnStart, dataRowStart, iDColumnEnd, dataRowEnd, relativeAbundanceColumn):
        self.iDColumnStart = iDColumnStart
        self.dataRowStart = dataRowStart
        self.iDColumnEnd = iDColumnEnd
        self.dataRowEnd = dataRowEnd
        self.relativeAbundanceColumn = relativeAbundanceColumn

    def AddMicroOrganismList(self, microOrganismList):
        self.MicroOrganisms = microOrganismList
        
    def FullDataSetInformation(self):
        return ("Name - " + self.sampleName +
                " | " + self.iDColumnStart + self.dataRowStart + ":" + self.iDColumnEnd + self.dataRowEnd +
                " | " + self.relativeAbundanceColumn + self.dataStart +
                ":" + self.relativeAbundanceColumn + self.dataRowEnd)

    def RelativeAbundanceRange(self):
        return (self.relativeAbundanceColumn + self.dataRowStart + ":" + self.relativeAbundanceColumn + self.dataRowEnd)

    def BacteriaNameRange(self):
        return (self.iDColumnStart + self.dataRowStart + ":" + self.iDColumnEnd + self.dataRowEnd)

    def ReturnRelativeAbundanceOfMicroOraganism(self, microOrganismIDToCheck_Hash):
        for microOrganism in self.MicroOrganisms:
            hashIDToCheck = microOrganism.hashID
            if(hashIDToCheck == microOrganismIDToCheck_Hash):
                return microOrganism.relativeAbundance
        #else
        return 0


    def CreateMicroOrganismList(self, BacteriaNameList, RelativeAbundanceList):
        #Takes a 1-Dimensional OpenPyxl List and converts it into a python list
        #Intialise the List that will be filled with python values
        microOrganismList = []

        #For every item in the openpyxl List, grab the value of that item and append it to the Python List
        for i in range(len(RelativeAbundanceList)):
            #Grab the relevant values at every row
            newDomain = BacteriaNameList[i][0].value
            newPhylum = BacteriaNameList[i][1].value
            newClassification = BacteriaNameList[i][2].value
            newOrder = BacteriaNameList[i][3].value
            newFamily = BacteriaNameList[i][4].value
            newGenus =BacteriaNameList[i][5].value
            newRelativeAbundance = RelativeAbundanceList[i][0].value

            #Create a new MicroOrganism object
            newMicroOrganism = MicroOrganism(newDomain,
                                             newPhylum,
                                             newClassification,
                                             newOrder,
                                             newFamily,
                                             newGenus,
                                             newRelativeAbundance)
            #Generate the Hash ID for the Microorganism
            newMicroOrganism.GenerateSelfHashID()

            #Add the new Microogranim to the list of microorganisms
            microOrganismList.append(newMicroOrganism)

        #Return list microOrganism's
        return microOrganismList  

    def DetermineMicroOrganisms(self):
        #Load the Data set's Workbook
        excelWorkBook = load_workbook(filename=self.workBookName, data_only=True)

        #Load the Data set's Sheet
        excelSheet = excelWorkBook[self.workSheetName]

        #Load the Data set's Relative Abundance
        openpyxlList_BacteriaName = excelSheet[self.BacteriaNameRange()]
        openpyxlList_RelativeAbundance = excelSheet[self.RelativeAbundanceRange()]

        #Create a microOrganism List
        microOrganismList = self.CreateMicroOrganismList(openpyxlList_BacteriaName, openpyxlList_RelativeAbundance)
            
        #Add the new Microbe List to the provided DataSet
        self.AddMicroOrganismList(microOrganismList)

    def GenerateTree(self, workBookName, minColumn, maxColumn, minRow, maxRow, functionDictionary, taxonDictionary):
        #Load the NoteBook
        excelWorkBook = load_workbook(filename=workBookName, read_only=True)
        excelWorkBook = excelWorkBook.active
        #Create a new Branch
        #Go through each row in the excel workbook
        breakIndex = 10000000
        currentIndex = 0
        for row in excelWorkBook.iter_rows():
            currentIndex += 1
            #If the sample name is the same as the given data set
            if row[0].value == self.sampleName:
                #Create new branch with the taxon ID from the parent file
                taxonID = row[2].value
                taxonRelativeAbundance = row[4].value
                newBranch = Branch(taxonID, taxonRelativeAbundance)

                #Convert taxon ID to micro organism
                newBranch.GenerateTaxonIDHashID()
                
                #Check if the branch for this taxonID has been created already
                branchExists = False
                for branch in self.microOrganismTree:
                    if branch.taxonIDHashID == newBranch.taxonIDHashID:
                        #Create new Function Description pair and generate hash ID
                        newFunctionDescriptionPair = FunctionDescriptionPair(row[1].value)
                        newFunctionDescriptionPair.GenerateFunctionHashID()
                        newFunctionDescriptionPair.GenerateDescriptionFromDictionary(functionDictionary)

                        #Add new pair to already existing branch
                        branch.AddNewFunctionDescriptionPair(newFunctionDescriptionPair)
                        branchExists = True
                        
                
                #If branch does not run this section
                if branchExists == False:
                    #Branch does not exist, so fill the branch with its leaf
                    #Generate Microoragnism from Taxon ID
                    newBranch.GenerateMicroOrgansimFromTaxonID(taxonDictionary)
                    
                    #Add Function Description pair and generate hash ID
                    newFunction = row[1].value
                    newFunctionDescriptionPair = FunctionDescriptionPair(newFunction)
                    newFunctionDescriptionPair.GenerateFunctionHashID()
                    newFunctionDescriptionPair.GenerateDescriptionFromDictionary(functionDictionary)

                    #Add new pair to new branch
                    newBranch.AddNewFunctionDescriptionPair(newFunctionDescriptionPair)

                    #Add new branch to tree
                    self.AddBranchToTree(newBranch)

            if currentIndex == breakIndex:
                return

    def AddBranchToTree(self, newBranch):                        
        #Cache temporary list
        tempList = self.microOrganismTree

        #Add new pair to temporary list
        tempList.append(newBranch)

        #Replace old list with new list
        self.microOrganismTree = tempList

    
class MicroOrganism:
    def __init__(self, domain, phylum, classification, order, family, genus, relativeAbundance):
        self.domain = domain
        self.phylum = phylum
        self.classification = classification
        self.order = order
        self.family = family
        self.genus = genus
        self.relativeAbundance = relativeAbundance
        self.hashID = None

    def FullBacteriaID(self):
        return (self.domain +";"+
                self.phylum +";"+
                self.classification +";"+
                self.order +";"+
                self.family +";"+
                self.genus)

    def HighestNameStructure(self):
        if(self.genus == "__") or (self.genus == "g__uncultured"):
            if(self.family == "__") or (self.family == "f__uncultured"):
                if(self.order == "__") or (self.order == "o__uncultured"):
                    if(self.classification == "__") or (self.classification == "c__uncultured"):
                        if(self.phylum == "__") or (self.phylum == "p__uncultured"):
                            return self.domain
                        else:
                            return self.phylum
                    else:
                        return self.classification
                else:
                    return self.order
            else:
                return self.family
        else:
            return self.genus
    
    def RelativeAbundance(self):
        return self.relativeAbundance

    def GenerateSelfHashID(self):
        self.hashID = hashlib.md5(self.FullBacteriaID().encode()).hexdigest()


class Branch:
    def __init__(self, taxonID, taxonRelativeAbundance):
        self.taxonID = taxonID
        self.taxonRelativeAbundance = taxonRelativeAbundance
        self.taxonIDHashID = None
        self.microOrganism = None
        self.functionDescriptionPairs = []

    def GenerateTaxonIDHashID(self):
        self.taxonIDHashID = hashlib.md5(self.taxonID.encode()).hexdigest()
        
    def GenerateMicroOrgansimFromTaxonID(self, chosenTaxonDictionary):
        #For every row in the dictionary
        for row in chosenTaxonDictionary.DictionaryData:
            #Compare the taxon ID of this branch with the taxon ID in each row
            if(self.taxonIDHashID == hashlib.md5(row[0].value.encode()).hexdigest()):
                #When it equals, then grab the taxon list
                taxonList = row[1].value.split("; ")
                
                #Iterate 6 times to make sure the new MO only has a length of 6, this takes it up to the same number as original document
                for i in range(6):
                    while True:
                        try:
                            taxonList[i]
                            break
                        except IndexError:
                            taxonList.append("__")

                #Create a new MicroOrganism object
                self.microOrganism = MicroOrganism(taxonList[0],
                                                     taxonList[1],
                                                     taxonList[2],
                                                     taxonList[3],
                                                     taxonList[4],
                                                     taxonList[5],
                                                     0)
                #Generate hash ID for itself
                self.microOrganism.GenerateSelfHashID()
                
     

    def AddNewFunctionDescriptionPair(self, newPair):            
        #Cache temporary list
        tempList = self.functionDescriptionPairs

        #Add new pair to temporary list
        tempList.append(newPair)

        #Replace old list with new list
        self.functionDescriptionPairs = tempList


class FunctionDescriptionPair:
    def __init__(self, function):
        self.function = function
        self.functionHashID = None
        self.description = None

    def GenerateFunctionHashID(self):
        self.functionHashID = hashlib.md5(self.function.encode()).hexdigest()

    def GenerateDescriptionFromDictionary(self, chosenFunctionDictionary):
        #Go through every row in the chosen dictionary
        for i in range(len(chosenFunctionDictionary.DictionaryData)):
            if (self.functionHashID == hashlib.md5(chosenFunctionDictionary.DictionaryData[i][0].value.encode()).hexdigest()):
                self.description = chosenFunctionDictionary.DictionaryData[i][1].value

        
class LookUpDictionary:
    def __init__(self, workBookName, workSheetName, columnStart, dataRowStart, columnEnd, dataRowEnd):
        self.workBookName = workBookName
        self.workSheetName = workSheetName
        self.columnStart = columnStart
        self.dataRowStart = dataRowStart
        self.columnEnd = columnEnd
        self.dataRowEnd = dataRowEnd
        self.DictionaryData = None
        
        self.GenerateDictionary()

    def DataRange(self):
        return (self.columnStart + self.dataRowStart + ":" + self.columnEnd + self.dataRowEnd)

    def GenerateDictionary(self):        
        #Load the Data set's Workbook
        excelWorkBook = load_workbook(filename=self.workBookName, data_only=True)

        #Load the Data set's Sheet
        excelSheet = excelWorkBook[self.workSheetName]

        #Load the Data set
        self.DictionaryData = excelSheet[self.DataRange()]
        
    
###Visualisation###   
def MakeHeatMap(dataSets, LinewidthOveride = 0, PercentToIgnore = 0):
    #Initialise the parts of the Data Frame, a list of RA for each DataSet,
    #   and a list of names for the data sets.
    RALists = []
    dataSetNameLists = []
    microOrganismNameList = []

    #Initialise cut down lists at the same time to make them the same size
    cutdownRALists = []
    cutdownMicroOrganismNameList = []

    #Populate lists
    for dataSet in dataSets:
        RALists.append([])
        cutdownRALists.append([])

        if(dataSet.nameOveride == None):
            dataSetNameLists.append(dataSet.sampleName)
        else:
             dataSetNameLists.append(dataSet.nameOveride)

    #Initialise a list of all microbacteria that appear (Note this uses hashTables for optimisation)
    #   Add a hashID to the HashID list to make searching faster, also add the highest order name ID.
    allMicrobacteriaPresent = []

    for dataSet in dataSets:
        for i in range(len(dataSet.MicroOrganisms)):
            hashIDToCheck = dataSet.MicroOrganisms[i].hashID
            if(hashIDToCheck not in allMicrobacteriaPresent):
                microOrganismNameList.append(dataSet.MicroOrganisms[i].HighestNameStructure())
                allMicrobacteriaPresent.append(hashIDToCheck)

    #Now go through the entire mO list and add the relative abundance of that mO of each data set
    #   If there is no entry for that data set then the value returned is 0
    
    for mO in allMicrobacteriaPresent:
        for i in range(len(dataSets)):
            RALists[i].append(dataSets[i].ReturnRelativeAbundanceOfMicroOraganism(mO))
            
    #You now have; a list of relative abundances for each data set (RALists)
    #              a list of microorganism names (microOrganismNameList)
    #              a list of dataSet names (dataSetNameLists)

    #Cut down the list to ignore any values that are below x%
    for i in range(len(microOrganismNameList)):
        includeMicroOrganism = False
        for RAList in RALists:
            if (RAList[i] > (PercentToIgnore/100)):
                includeMicroOrganism = True

        if (includeMicroOrganism):
            cutdownMicroOrganismNameList.append(microOrganismNameList[i])
            for j in range(len(RALists)):
                if(RALists[j][i] > (PercentToIgnore/100)):
                    cutdownRALists[j].append(RALists[j][i] * 100)
                else:
                    cutdownRALists[j].append(0)

    #Now Create a pandas frame work to hold all this information
    pandaFrameWork = pd.DataFrame(cutdownRALists, columns=cutdownMicroOrganismNameList, index=dataSetNameLists)
    pandaFrameWork_transpose = pandaFrameWork.transpose()

    #Plot Data
    HeatMapPlot(pandaFrameWork_transpose, LinewidthOveride)

def HeatMapPlot(data, linewidthOveride):
    sns.heatmap(data, annot=True, linewidths=linewidthOveride, cmap="YlGnBu", yticklabels=True)
    plt.tight_layout()
    plt.show()


def DrawTreeDiagram(branch):
    g = graphviz.Graph('G', engine='sfdp')
    for fdPair in branch.functionDescriptionPairs:
        g.edge(fdPair.function, branch.microOrganism.FullBacteriaID())


    g.view()



###INITIALISE DATASETS###
dataSet1 = SampleInformation("HannahExcelData.xlsx", "Sheet1", "AA1")
dataSet2 = SampleInformation("HannahExcelData.xlsx", "Sheet1", "BB2")
dataSet3 = SampleInformation("HannahExcelData.xlsx", "Sheet1", "CC01B")

dataSet1.nameOveride =  "Hannah1"
dataSet2.nameOveride =  "Hannah2"


###SET UP DATA SET
dataSet1.SetUpMicroOrganismDatSet( "A", "3", "F", "203", "G")
dataSet2.SetUpMicroOrganismDatSet( "A", "3", "F", "203", "G")
dataSet3.SetUpMicroOrganismDatSet( "A", "3", "F", "203", "G")


###DETERMINE MICRO-ORAGANISMS###
dataSet1.DetermineMicroOrganisms()
dataSet2.DetermineMicroOrganisms()
dataSet3.DetermineMicroOrganisms()


###Plot HeatMap of MICRO-ORGANISMS###
MakeHeatMap([dataSet1, dataSet2], LinewidthOveride = 0, PercentToIgnore = 1)

###Set Up Dictionaries
FunctionDictionary = LookUpDictionary("Function Dictionary.xlsx", "Sheet 1 - path_abun_unstrat_des", "A", "3", "B", "404")
TaxonDictionary = LookUpDictionary("Taxon dictionary.xlsx", "Sheet 1 - metadata (9)", "A", "4", "C", "2672")

###SET UP TAXON and DESCRIPTION DATA
dataSet3.GenerateTree("Parent.xlsx", "A", "H", "1", "935127", FunctionDictionary, TaxonDictionary)

###Visualise Tree Diagram
DrawTreeDiagram(dataSet3.microOrganismTree[0])
